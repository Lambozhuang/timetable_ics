import openpyxl
import datetime
import time

# columnOfClassName = 0
# columnOfStartTime = 1
# columnOfEndTime = 2
# columnOfWeekDay = 3
# columnOfStartWeek = 4
# columnOfEndWeek = 5
# columnOfLocation = 6
# columnOfStartDate = 8
# _columnOfStartTime = 11
# _columnOfEndTime = 12

columnOfClassName = 1
columnOfStartTime = 2
columnOfEndTime = 3
columnOfWeekDay = 4
columnOfStartWeek = 5
columnOfEndWeek = 6
columnOfLocation = 7
columnOfStartDate = 9
_columnOfStartTime = 12
_columnOfEndTime = 13

# main

outputStr = 'BEGIN:VCALENDAR\nMETHOD:PUBLISH\nVERSION:2.0\nX-WR-CALNAME:课表\nPRODID:-//Apple Inc.//macOS 11.2.2//EN\nX-APPLE-CALENDAR-COLOR:#711A76\nX-WR-TIMEZONE:Asia/Shanghai\nCALSCALE:GREGORIAN\nBEGIN:VTIMEZONE\nTZID:Asia/Shanghai\nBEGIN:STANDARD\nTZOFFSETFROM:+0900\nRRULE:FREQ=YEARLY;UNTIL=19910914T170000Z;BYMONTH=9;BYDAY=3SU\nDTSTART:19890917T020000\nTZNAME:GMT+8\nTZOFFSETTO:+0800\nEND:STANDARD\nBEGIN:DAYLIGHT\nTZOFFSETFROM:+0800\nDTSTART:19910414T020000\nTZNAME:GMT+8\nTZOFFSETTO:+0900\nRDATE:19910414T020000\nEND:DAYLIGHT\nEND:VTIMEZONE\n'

timeMap = {}
timeSeq = ('1', '2', '3', '4', '5', '6', '7', '8', '9', '10', '11', '12')
timeMap = timeMap.fromkeys(timeSeq)
startWeekDate = ''
numberOfClass = 0

# 打开excel工作簿
data = openpyxl.load_workbook('timetable.xlsx')
# table = data.sheets()[0]
table = data.active

# 读取学期第一周星期一开始日期
startWeekDate = str(table.cell(2, columnOfStartDate).value)
startWeekDate = datetime.datetime.strptime(startWeekDate, '%Y%m%d')

print(table.max_column)
print(table.max_row)
print(startWeekDate)

# 读取课程数量和节数对应时间
for i in range(2, table.max_row + 1):
  tempList = []
  tempList.append(table.cell(i, _columnOfStartTime).value)
  tempList.append(table.cell(i, _columnOfEndTime).value)
  # print(tempList)
  if i < 14:
    timeMap[str(i - 1)] = tempList
  if table.cell(i, columnOfClassName).value != None:
    numberOfClass += 1

print(numberOfClass)
  
for i in range(2, numberOfClass + 2):
  startWeek = int(table.cell(i, columnOfStartWeek).value)
  endWeek = int(table.cell(i, columnOfEndWeek).value)
  numberOfWeek = endWeek - startWeek + 1
  weekDay = int(table.cell(i, columnOfWeekDay).value)
  location = table.cell(i, columnOfLocation).value
  summary = table.cell(i, columnOfClassName).value
  startTime = table.cell(i, columnOfStartTime).value
  endTime = table.cell(i, columnOfEndTime).value
  startDate = startWeekDate + datetime.timedelta((startWeek - 1) * 7 + weekDay - 1)
  date = startDate
  eventStr = ''
  if location == None:
    location = ""

  print(startWeek)

  for j in range(0, numberOfWeek):
    temp = 'BEGIN:VEVENT\nTRANSP:OPAQUE\n'
    temp += 'DTSTART;TZID=Asia/Shanghai:'
    temp = temp + date.strftime('%Y%m%d') + 'T' + timeMap[str(startTime)][0] + '00\n'
    temp += 'DTEND;TZID=Asia/Shanghai:'
    temp = temp + date.strftime('%Y%m%d') + 'T' + timeMap[str(endTime)][1] + '00\n'
    temp += 'LOCATION:'
    temp = temp + location + '\n'
    temp += 'SUMMARY:'
    temp = temp + summary + '\n'
    temp += 'BEGIN:VALARM\nTRIGGER:-PT15M\nDESCRIPTION:提醒事项\nACTION:DISPLAY\nEND:VALARM\nEND:VEVENT\n'
    eventStr += temp
    temp = ''
    date = date + datetime.timedelta(7)

  outputStr += eventStr

print(outputStr)
file = open('timetable.ics', 'wb')
file.write(outputStr.encode('utf-8'))
file.close()
